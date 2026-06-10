"""Importador da planilha de Negociação da B3 (Área do Investidor).

Lê o Excel (.xlsx) ou CSV exportado em Extratos > Negociação e devolve negócios
no formato interno. Casa as colunas de forma flexível (sem acento, minúsculas):

    Data do Negócio | Tipo de Movimentação | Mercado | Vencimento |
    Instituição | Código de Negociação | Quantidade | Preço | Valor

Observações importantes:
- A B3 informa o "Valor" já em REAIS (inclusive para futuros, onde "Preço" está
  em pontos). Por isso usamos preço = Valor / Quantidade, deixando tudo em reais
  e mantendo a apuração correta para ações E futuros.
- O extrato de negociação NÃO traz custos (corretagem/emolumentos) nem IRRF —
  as notas importadas entram com custo zero (aproximação; ajustável depois).
"""
from __future__ import annotations

import csv
import io
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation


def _norm(s) -> str:
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.strip().lower()


def _market(raw) -> str:
    r = _norm(raw)
    if "fracion" in r:
        return "FRACIONARIO"
    if "futur" in r:
        return "FUTURO"
    if "opca" in r or "opcao" in r or "opç" in r:
        return "OPCAO"
    if "termo" in r:
        return "TERMO"
    return "VISTA"


def _to_decimal(v) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, int | float):
        return Decimal(str(v))
    t = str(v).strip().replace("R$", "").strip()
    # pt-BR: 1.234,56 -> 1234.56
    if "," in t:
        t = t.replace(".", "").replace(",", ".")
    t = re.sub(r"[^0-9.\-]", "", t)
    try:
        return Decimal(t) if t else Decimal("0")
    except InvalidOperation:
        return Decimal("0")


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _match_columns(header: list) -> dict:
    """Mapeia os índices das colunas que interessam a partir do cabeçalho."""
    cols = {}
    for i, h in enumerate(header):
        n = _norm(h)
        if "data" in n and "date" not in cols:
            cols["date"] = i
        elif "moviment" in n and "side" not in cols:   # Tipo de Movimentação
            cols["side"] = i
        elif n == "mercado" and "market" not in cols:
            cols["market"] = i
        elif "institu" in n and "broker" not in cols:
            cols["broker"] = i
        elif "cod" in n and "asset" not in cols:        # Código de Negociação
            cols["asset"] = i
        elif "quantidade" in n and "qty" not in cols:
            cols["qty"] = i
        elif ("preco" in n or "preço" in n) and "price" not in cols:
            cols["price"] = i
        elif "valor" in n and "value" not in cols:
            cols["value"] = i
    return cols


def _rows_from_xlsx(stream) -> list:
    from openpyxl import load_workbook
    wb = load_workbook(stream, read_only=True, data_only=True)
    ws = wb.active
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _rows_from_csv(stream) -> list:
    raw = stream.read()
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8-sig", errors="replace")
    sample = raw[:2000]
    delim = ";" if sample.count(";") >= sample.count(",") else ","
    return [list(r) for r in csv.reader(io.StringIO(raw), delimiter=delim)]


def parse(stream, filename: str) -> dict:
    """Lê a planilha e devolve {'trades': [...], 'warnings': [...]}.

    Cada trade: dict(trade_date, asset, market, side, quantity, price, gross_value, broker).
    """
    name = (filename or "").lower()
    rows = _rows_from_xlsx(stream) if name.endswith(".xlsx") else _rows_from_csv(stream)

    # Encontra a linha de cabeçalho (a que tem 'quantidade' e 'movimenta'/'mercado').
    header_idx, cols = None, {}
    for i, row in enumerate(rows[:15]):
        c = _match_columns(row)
        if "qty" in c and "asset" in c and ("side" in c or "market" in c):
            header_idx, cols = i, c
            break
    if header_idx is None:
        return {"trades": [], "warnings": [
            "Não reconheci o cabeçalho da planilha da B3. Confira se é o extrato "
            "de Negociação (Extratos > Negociação)."]}

    def cell(row, key):
        i = cols.get(key)
        return row[i] if i is not None and i < len(row) else None

    trades, warnings, ignoradas = [], [], 0
    for row in rows[header_idx + 1:]:
        if not row or all(c in (None, "") for c in row):
            continue
        side_raw = _norm(cell(row, "side"))
        if side_raw.startswith("compra"):
            side = "C"
        elif side_raw.startswith("venda"):
            side = "V"
        else:
            ignoradas += 1
            continue

        d = _to_date(cell(row, "date"))
        qty = _to_decimal(cell(row, "qty"))
        value = _to_decimal(cell(row, "value"))
        asset = str(cell(row, "asset") or "").strip().upper()
        if not asset or qty <= 0 or d is None:
            ignoradas += 1
            continue
        price = (value / qty) if qty else _to_decimal(cell(row, "price"))

        trades.append({
            "trade_date": d, "asset": asset, "market": _market(cell(row, "market")),
            "side": side, "quantity": qty, "price": price,
            "gross_value": value if value > 0 else qty * price,
            "broker": str(cell(row, "broker") or "B3").strip()[:60] or "B3",
        })

    if not trades:
        warnings.append("Nenhuma operação de compra/venda reconhecida na planilha.")
    elif ignoradas:
        warnings.append(f"{ignoradas} linha(s) ignorada(s) (sem compra/venda válida).")
    return {"trades": trades, "warnings": warnings}


# --------------------------------------------------------------------------- #
# Reconciliação: planilha da B3 × notas importadas no app
# --------------------------------------------------------------------------- #
def reconcile(b3_trades: list, app_trades: list) -> dict:
    """Compara os negócios da planilha B3 com os já registrados no app.

    Agrega por (data, ativo, lado) e aponta:
      - only_b3 : está na B3 e NÃO está no app (falta nota!)
      - only_app: está no app e NÃO está na B3 (lançamento manual errado?)
      - mismatch: existe nos dois, mas quantidade ou financeiro divergem
    Financeiro tolera 1% (a B3 arredonda; notas têm taxas).
    """
    def agg(items, get):
        out = {}
        for t in items:
            d, asset, side, qty, gross = get(t)
            key = (d, asset.upper().strip(), side)
            cur = out.setdefault(key, {"qty": Decimal("0"), "gross": Decimal("0")})
            cur["qty"] += qty
            cur["gross"] += gross
        return out

    b3 = agg(b3_trades, lambda t: (t["trade_date"], t["asset"], t["side"],
                                   Decimal(str(t["quantity"])),
                                   Decimal(str(t["gross_value"]))))
    app = agg(app_trades, lambda t: (t.trade_date, t.asset, t.side,
                                     Decimal(str(t.quantity)),
                                     Decimal(str(t.gross_value))))

    only_b3, only_app, mismatch, matched = [], [], [], 0
    for key in sorted(set(b3) | set(app)):
        d, asset, side = key
        row = {"date": d, "asset": asset, "side": side}
        if key not in app:
            only_b3.append({**row, **b3[key]})
        elif key not in b3:
            only_app.append({**row, **app[key]})
        else:
            qa, qb = app[key]["qty"], b3[key]["qty"]
            ga, gb = app[key]["gross"], b3[key]["gross"]
            tol = max(abs(gb) * Decimal("0.01"), Decimal("0.05"))
            if qa != qb or abs(ga - gb) > tol:
                mismatch.append({**row, "app_qty": qa, "b3_qty": qb,
                                 "app_gross": ga, "b3_gross": gb})
            else:
                matched += 1
    return {"only_b3": only_b3, "only_app": only_app,
            "mismatch": mismatch, "matched": matched,
            "total_keys": len(set(b3) | set(app))}
